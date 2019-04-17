"""TODO."""

from openpyxl import load_workbook

from cahpy.list import group_list_to_dict


def worksheet_to_dict(filename, worksheetName, keyFieldname, sortKey=None):
    """TODO."""
    return group_list_to_dict(
        worksheet_to_list_of_dicts(filename, worksheetName), keyFieldname)


def worksheet_to_list_of_dicts(filename, worksheetName):
    """TODO."""
    worksheet = load_workbook(filename=str(filename),
                              read_only=True)[worksheetName]
    headers = [cell.value
               for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]

    return [{
        header: cell.value
        for header, cell in zip(headers, row)
    } for row in worksheet.iter_rows(min_row=2)]
